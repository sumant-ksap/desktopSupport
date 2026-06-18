"""
Generate a formatted Excel decision chart from the email analysis results.
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"
ROW_ODD     = "EBF0FA"
ROW_EVEN    = "FFFFFF"

IMPORTANCE_COLOURS = {
    "critical": "FF0000",
    "high":     "FF6600",
    "medium":   "FFD700",
    "low":      "00B050",
}

CATEGORY_LABELS = {
    "product_complaint":      "Product Complaint",
    "billing_issue":          "Billing Issue",
    "technical_support":      "Technical Support",
    "feature_request":        "Feature Request",
    "general_inquiry":        "General Inquiry",
    "newsletter_or_promo":    "Newsletter / Promo",
    "spam":                   "Spam",
    "internal_communication": "Internal",
    "order_status":           "Order Status",
    "feedback":               "Feedback",
    "other":                  "Other",
}

COLUMNS = [
    ("No.",             8),
    ("From",           30),
    ("Subject",        45),
    ("Date Received",  22),
    ("Category",       24),
    ("Importance",     14),
    ("Summary",        55),
    ("Suggested Action", 40),
    ("AI Confidence",  14),
    ("Action Taken",   20),
]


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HEADER_BG)


def _row_fill(row_idx: int) -> PatternFill:
    colour = ROW_ODD if row_idx % 2 == 0 else ROW_EVEN
    return PatternFill("solid", fgColor=colour)


def _importance_fill(level: str) -> PatternFill:
    colour = IMPORTANCE_COLOURS.get(level.lower(), "CCCCCC")
    return PatternFill("solid", fgColor=colour)


def build_excel(analysed_emails: list[dict], output_dir: Path) -> Path:
    """
    Build the decision-chart Excel workbook and return the file path.

    analysed_emails: list of dicts with keys:
        uid, subject, sender, date, body (from email_handler)
        + category, importance, is_product_complaint, summary,
          suggested_action, confidence (from ai_analyzer)
        + action_taken (set by main.py, e.g. "Forwarded as complaint")
    """
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, analysed_emails)

    # ── Detail sheet ───────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet("Decision Chart")
    _build_detail_sheet(ws_detail, analysed_emails)

    # ── Category breakdown sheet ───────────────────────────────────────────────
    ws_cat = wb.create_sheet("Category Breakdown")
    _build_category_sheet(ws_cat, analysed_emails)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"email_decision_chart_{timestamp}.xlsx"
    wb.save(out_path)
    print(f"[Excel] Report saved: {out_path}")
    return out_path


# ── Summary sheet ──────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, emails: list[dict]) -> None:
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    title_font = Font(name="Calibri", size=16, bold=True, color=HEADER_BG)
    label_font = Font(name="Calibri", size=11, bold=True)
    value_font = Font(name="Calibri", size=11)

    ws["A1"] = "Email Decision Chart — Executive Summary"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.merge_cells("A2:B2")

    total = len(emails)
    complaints = sum(1 for e in emails if e.get("is_product_complaint"))
    forwarded  = sum(1 for e in emails if e.get("action_taken") == "Forwarded as complaint")

    importance_counts = {}
    for e in emails:
        lvl = e.get("importance", "medium")
        importance_counts[lvl] = importance_counts.get(lvl, 0) + 1

    rows = [
        ("", ""),
        ("Metric", "Value"),
        ("Total Emails Processed", total),
        ("Product Complaints Found", complaints),
        ("Complaints Forwarded",    forwarded),
        ("Non-Complaint Emails",    total - complaints),
        ("", ""),
        ("Importance Breakdown", "Count"),
        ("Critical", importance_counts.get("critical", 0)),
        ("High",     importance_counts.get("high",     0)),
        ("Medium",   importance_counts.get("medium",   0)),
        ("Low",      importance_counts.get("low",      0)),
    ]

    for r_idx, (label, value) in enumerate(rows, start=3):
        cell_a = ws.cell(row=r_idx, column=1, value=label)
        cell_b = ws.cell(row=r_idx, column=2, value=value)
        if label in ("Metric", "Importance Breakdown"):
            for c in (cell_a, cell_b):
                c.font = Font(name="Calibri", bold=True, color="FFFFFF")
                c.fill = _header_fill()
        elif label:
            cell_a.font = label_font
            cell_b.font = value_font
            if label in ("Critical", "High", "Medium", "Low"):
                cell_b.fill = _importance_fill(label.lower())
        for c in (cell_a, cell_b):
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center")


# ── Detail / Decision Chart sheet ─────────────────────────────────────────────

def _build_detail_sheet(ws, emails: list[dict]) -> None:
    # Header row
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, em in enumerate(emails, start=2):
        fill = _row_fill(row_idx)
        values = [
            row_idx - 1,
            em.get("sender", ""),
            em.get("subject", ""),
            em.get("date", ""),
            CATEGORY_LABELS.get(em.get("category", "other"), em.get("category", "")),
            em.get("importance", "").capitalize(),
            em.get("summary", ""),
            em.get("suggested_action", ""),
            em.get("confidence", "").capitalize(),
            em.get("action_taken", "Included in report"),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _thin_border()
            if col_idx == 6:  # Importance column — colour-code
                cell.fill = _importance_fill(em.get("importance", ""))
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            else:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 40


# ── Category breakdown sheet ──────────────────────────────────────────────────

def _build_category_sheet(ws, emails: list[dict]) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    headers = ["Category", "Count", "% of Total"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color=HEADER_FG)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    category_counts: dict[str, int] = {}
    for em in emails:
        cat = em.get("category", "other")
        category_counts[cat] = category_counts.get(cat, 0) + 1

    total = len(emails) or 1
    sorted_cats = sorted(category_counts.items(), key=lambda x: -x[1])

    for row_idx, (cat, count) in enumerate(sorted_cats, start=2):
        fill = _row_fill(row_idx)
        label = CATEGORY_LABELS.get(cat, cat.replace("_", " ").title())
        pct = round(count / total * 100, 1)
        for col_idx, val in enumerate([label, count, f"{pct}%"], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")

    # Totals row
    total_row = len(sorted_cats) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value="100%").font = Font(bold=True)
    for col_idx in range(1, 4):
        ws.cell(row=total_row, column=col_idx).border = _thin_border()
